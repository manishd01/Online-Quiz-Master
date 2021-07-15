from django.shortcuts import render
from django.http import HttpResponse, HttpResponseRedirect
import os
from .models import *
# Create your views here.
# count =1
def homepage(request):
    # return HttpResponse('hello')
    return render(request,'homepage.html')

def take_test(request):
    return render(request,'taketest.html')

def about_us(request):
    return render(request,'about_us.html')

def contact_us(request):
    return render(request,'contact_us.html')

def start_test(request):
    subject=request.POST['subject']
    num_of_ques=request.POST['num_of_ques']
    candidate_name=request.POST['candidate_name']
    ok=details_test(subject=subject,num_of_ques=num_of_ques,candidate_name=candidate_name)
    ok.save()
    return HttpResponseRedirect('http://127.0.0.1:8000/take_test/start_test/questionPaper/')

def questionPaper(request):
    print('questionPaper function')
    import openpyxl
    import random
    # wyb=openpyxl.load_workbook("C:\\pandas_numpy\\testx.xlsx")
    get_subject=details_test().show_record()
    get_subject=get_subject[1]
    print(type(get_subject))
    wb_obj=openpyxl.load_workbook("C:\\onlineCompititiveExaminationSystem\\questionBank\\all_sub.xlsx")
    sh_obj=wb_obj[get_subject]
    ques=random.randint(2,51)
    # print(sh_obj.cell(row=ques,column=7.value)
    ques_num=sh_obj.cell(row=ques,column=1).value
    ques_statement=sh_obj.cell(row=ques,column=2).value
    opt1=sh_obj.cell(row=ques,column=3).value
    opt2=sh_obj.cell(row=ques,column=4).value
    opt3=sh_obj.cell(row=ques,column=5).value
    opt4=sh_obj.cell(row=ques,column=6).value
    ans=sh_obj.cell(row=ques,column=7).value
    answ=open('readans.txt','w+')
    answ.write(str(ans))
    answ.close()

    # def ansf():
    #     return ans
    # ans=int(ans) #not wrking because we can  not covert 'd' into int but we convert '6' into int with int()
    # ans=ord(ans)  #working fine
    # print(ans)
    # acess_ansVal=sh_obj.cell(row=ques,column=ans-64+2).value #ord() get string char and get int value as per ASCII table
    #ans=(value i.e A B C or D ) ,-64charters as per ASCII and +2 column because A B C D option starts from col3
    # print(acess_ansVal)
    values_pass={'ques_num':ques_num,'ques_statement':ques_statement,'opt1':opt1,'opt2':opt2,
    'opt3':opt3,'opt4':opt4,'ans':answ}
    return render(request,'questionPaper.html',values_pass)
def check_que_num(request):
    print('check_que_num function')
    # from .views import questionPaper
    # hel=details_test.objects.all()# hel=details_test.objects.latest('num_of_ques')# hel=details_test.objects.earliest('num_of_ques')
    NOQ_views=details_test().show_record()  #NOQ-> NumberOfQuestions
    NOQ_views1=NOQ_views[2]
    print(NOQ_views)
    print(type(NOQ_views))
    NOQ_views=int(NOQ_views1)
    print(NOQ_views1)
    print(type(NOQ_views))  # for coun in range(1,NOQ_views):# for i in range(NOQ_views)#     print(count)#     i=0#     i=count+1#     # global count #     count = i# global nuk
    fr=open('file1.txt','r+')
    i=int(fr.read())
    print(i)
    fr.close()

    entered_ans=details_answers().show_entered_ans()
    print(entered_ans)
    entered_ans=entered_ans[1]
    score_re=open('score.txt','r+')
    marks=int(score_re.read())
    score_re.close()
    # ans=questionPaper(request).ansf()   # ans=global(ansk) # 
    # ans=questionPaper(request) 
    # ans=ans1[1]  # ans=ans  # print(answ)
    Act_ans=open('readans.txt','r+')
    ans=Act_ans.read()
    Act_ans.close()
    # print(str(ans))
    if entered_ans==ans:
        # marks=int(marks)+1
        marks+=1
        print(" changed_marks: ",marks)
        score_wri=open('score.txt','w+')
        score_wri.write(str(marks))
        score_wri.close()
    else:
        score_wri=open('score.txt','w+')
        score_wri.write(str(marks))
        print("unchanges marks:",marks)
        score_wri.close()
    while(i<NOQ_views):
        # print()
        print(i)
        i+=1
        fw=open('file1.txt','w+')
        fw.write(str(i))
        fw.close()  
        return HttpResponseRedirect('http://127.0.0.1:8000/take_test/start_test/questionPaper/')
    else:
        fw2=open('file1.txt','w+')
        fw2.write('1')
        fw2.close()
        score_wri=open('score.txt','w+')
        score_wri.write('0')
        # print("unchanges marks:",marks)
        score_wri.close()
        # print(NOQ_views[3])'
        NOQ_views=details_test().show_record()
        name=NOQ_views[3]
        return render(request,'result.html',{'marks':marks,'name': name})

def save_answers(request):
    print('save_answers_function')
    option_selected=request.POST['option_selected']
    t=details_answers(option_selected=option_selected)
    t.save()
    # print(details_answers.option_selected.values_list().last()) #objects.values_list().last() 
    
    return HttpResponseRedirect('http://127.0.0.1:8000/take_test/start_test/questionPaper/check_queN/')


#questionpaper -> saveanswers -> checkquenum -> quesitonpaper